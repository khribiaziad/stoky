"""
One-time (and reusable) Excel import for StreetStore-style order sheets.
POST /platform/stores/{store_id}/import-excel
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from database import get_db
from auth import get_current_user
import models, re, io
from datetime import datetime

router = APIRouter(prefix="/platform", tags=["platform"])


def require_super_admin(user: models.User = Depends(get_current_user)):
    if user.role != "super_admin":
        raise HTTPException(status_code=403, detail="Super admin access required")
    return user


# ── Product catalog (scraped from streetstore.store) ─────────────────────────

PRODUCTS = [
    {"name": "jorts",                                      "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "Gray baggy",                                 "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "Baggy jeans",                                "price": 179, "sizes": ["36","38","40","42","44","46","48","50","52","54"]},
    {"name": "Wide Leg jeans",                             "price": 179, "sizes": ["36","38","40","42","44"], "colors": ["Blue","Black"]},
    {"name": "Charlie",                                    "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "Ensemble Jeans 2 Piece",                     "price": 299, "sizes": ["36","38","40","42","44"]},
    {"name": "Durty baggy jeans",                          "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "Dark Blue Denim Jacket & Wide-Leg Pants Set","price": 299, "sizes": ["38","40","42","44"]},
    {"name": "Light-wash blue jeans",                      "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "Baggy Wide Leg Denim Jeans",                 "price": 179, "sizes": ["36","38","40","42","44","46","48","50","52","54"]},
    {"name": "Jean Skirts",                                "price": 179, "sizes": ["36","38","40","42","44","46","48","50","52","54"]},
    {"name": "Dark blue Jeans",                            "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "High-Rise Dark Blue Jeans",                  "price": 179, "sizes": ["36","38","40","42","44","46","48","50","52","54"]},
    {"name": "Patte d'éléphant jean",                      "price": 179, "sizes": ["36","38","40","42","44"]},
    {"name": "Brown wide-leg jean",                        "price": 179, "sizes": ["36","38","40","42","44"]},
]

# Keywords → canonical product name (longest/most specific first)
PRODUCT_KEYWORDS = [
    ("dark blue denim jacket",          "Dark Blue Denim Jacket & Wide-Leg Pants Set"),
    ("denim jacket",                    "Dark Blue Denim Jacket & Wide-Leg Pants Set"),
    ("baggy wide leg denim",            "Baggy Wide Leg Denim Jeans"),
    ("baggy wide leg",                  "Baggy Wide Leg Denim Jeans"),
    ("high-rise dark blue",             "High-Rise Dark Blue Jeans"),
    ("high rise dark blue",             "High-Rise Dark Blue Jeans"),
    ("high-rise",                       "High-Rise Dark Blue Jeans"),
    ("light-wash blue",                 "Light-wash blue jeans"),
    ("light wash blue",                 "Light-wash blue jeans"),
    ("light-wash",                      "Light-wash blue jeans"),
    ("light wash",                      "Light-wash blue jeans"),
    ("dark blue jeans",                 "Dark blue Jeans"),
    ("dark blue jean",                  "Dark blue Jeans"),
    ("brown wide",                      "Brown wide-leg jean"),
    ("durty baggy",                     "Durty baggy jeans"),
    ("dirty baggy",                     "Durty baggy jeans"),
    ("black wide leg",                  "Wide Leg jeans"),
    ("wide leg jeans",                  "Wide Leg jeans"),
    ("wide leg jean",                   "Wide Leg jeans"),
    ("balloni",                         "Wide Leg jeans"),
    ("gray baggy",                      "Gray baggy"),
    ("grey baggy",                      "Gray baggy"),
    ("gray bag",                        "Gray baggy"),
    ("ensemble jeans",                  "Ensemble Jeans 2 Piece"),
    ("ensemble jean",                   "Ensemble Jeans 2 Piece"),
    ("2 piece",                         "Ensemble Jeans 2 Piece"),
    ("patte",                           "Patte d'éléphant jean"),
    ("elephant",                        "Patte d'éléphant jean"),
    ("charlie",                         "Charlie"),
    ("jean skirt",                      "Jean Skirts"),
    ("skirt",                           "Jean Skirts"),
    ("jorts",                           "jorts"),
    ("baggy jeans",                     "Baggy jeans"),
    ("baggy jean",                      "Baggy jeans"),
    ("baggy",                           "Baggy jeans"),
]


def match_product(raw_name: str) -> str | None:
    n = raw_name.lower().strip()
    for keyword, canonical in PRODUCT_KEYWORDS:
        if keyword in n:
            return canonical
    return None


def parse_price(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = re.sub(r'[^\d.]', '', str(val))
    try:
        return float(cleaned)
    except:
        return 0.0


def parse_size(val, raw_name: str = "") -> str | None:
    if val is not None:
        try:
            return str(int(float(val)))
        except:
            pass
    # Try extracting from name: trailing /38 or space 38
    m = re.search(r'[\s/](\d{2})(?:[\s/.]|$)', str(raw_name))
    return m.group(1) if m else None


def map_status(confirmation: str, delivery_status: str, has_tracking: bool):
    """Map Excel statuses → Stocky order status."""
    ds = (delivery_status or "").upper()
    conf = (confirmation or "").strip()

    if ds == "DELIVERED":
        return "delivered"
    if ds in ("CANCELED", "CANCELLED"):
        return "cancelled"
    if conf in ("Annulé", "Fake order", "Pas sérieux", "Comande en double"):
        return "cancelled"
    return "pending"


def map_delivery_status(raw: str) -> str:
    mapping = {
        "CONFIRMED": "Envoyé",
        "ENROUTE":   "En Route",
        "PICKUP":    "Ramassage",
        "DELIVERED": "DELIVERED",
        "TRANSIT":   "En Transit",
        "CANCELED":  "CANCELED",
        "REPORTED":  "REPORTED",
    }
    return mapping.get((raw or "").upper(), raw or "Envoyé")


@router.post("/stores/{store_id}/import-excel")
async def import_excel(
    store_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: models.User = Depends(require_super_admin),
):
    store = db.query(models.User).filter(
        models.User.id == store_id, models.User.role == "admin"
    ).first()
    if not store:
        raise HTTPException(404, "Store not found")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on server")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

    # Find the orders sheet (first sheet with "Commandes" or just first sheet)
    sheet_name = wb.sheetnames[0]
    for name in wb.sheetnames:
        if "commande" in name.lower() or "order" in name.lower():
            sheet_name = name
            break
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    # Find header row (contains "Order ID" or "Name")
    header_idx = 0
    for i, row in enumerate(rows):
        if row and any(str(v or "").strip().lower() in ("order id", "name", "nom") for v in row):
            header_idx = i
            break
    data_rows = [r for r in rows[header_idx + 1:] if any(v is not None for v in r)]

    # ── Step 1: Create products if they don't exist ───────────────────────────
    existing_products = {
        p.name: p for p in db.query(models.Product).filter_by(user_id=store_id).all()
    }
    created_products = {}

    for pd in PRODUCTS:
        name = pd["name"]
        if name in existing_products:
            created_products[name] = existing_products[name]
            continue
        product = models.Product(
            name=name,
            user_id=store_id,
            under_1kg=False,
        )
        db.add(product)
        db.flush()

        for size in pd["sizes"]:
            colors = pd.get("colors") or [None]
            for color in colors:
                variant = models.Variant(
                    product_id=product.id,
                    size=size,
                    color=color,
                    buying_price=85.0,
                    selling_price=float(pd["price"]),
                    stock=0,
                )
                db.add(variant)

        created_products[name] = product

    db.flush()

    # ── Step 2: Import orders ─────────────────────────────────────────────────
    existing_caleo_ids = {
        o.caleo_id for o in db.query(models.Order.caleo_id).filter_by(user_id=store_id).all()
    }

    created = 0
    skipped = 0
    unmatched_products = set()

    for row in data_rows:
        # Column indices (0-based): matches the confirmed sheet structure
        order_id    = row[0]
        order_date  = row[1]
        cust_name   = row[4]
        phone       = row[5]
        city        = row[7] or row[6]   # City API or City
        address     = row[8]
        raw_product = row[9]
        size_val    = row[11]
        color_val   = row[12]
        price_val   = row[13]
        quantity    = row[14]
        notes_val   = row[15]
        confirmation= row[16]
        delivery_co = row[20]
        delivery_st = row[22]
        tracking    = row[23]
        order_date2 = row[27] if len(row) > 27 else None

        if not cust_name and not phone:
            continue

        caleo_id = str(int(order_id)) if order_id and isinstance(order_id, float) else str(order_id or "")
        if caleo_id in existing_caleo_ids:
            skipped += 1
            continue

        date = order_date if isinstance(order_date, datetime) else (
            order_date2 if isinstance(order_date2, datetime) else datetime.now()
        )

        total = parse_price(price_val)
        qty = int(quantity) if quantity else 1
        if total > 0 and qty > 1:
            total = total  # total_amount is total price, not per-unit

        status = map_status(confirmation, delivery_st, bool(tracking))
        dl_status = map_delivery_status(delivery_st) if delivery_st else ("Envoyé" if tracking else None)
        provider = "olivraison" if "olivrais" in str(delivery_co or "").lower() else (
                   "forcelog"   if "forcelog"  in str(delivery_co or "").lower() else None)

        order = models.Order(
            user_id=store_id,
            caleo_id=caleo_id,
            order_date=date,
            customer_name=str(cust_name or "").strip(),
            customer_phone=str(phone or "").strip(),
            city=str(city or "").strip(),
            customer_address=str(address or "").strip(),
            total_amount=total,
            notes=str(notes_val or "").strip() or None,
            status=status,
            tracking_id=str(tracking).strip() if tracking else None,
            delivery_status=dl_status,
            delivery_provider=provider,
        )
        db.add(order)
        db.flush()

        # Order item
        raw_name = str(raw_product or "").strip()
        canonical = match_product(raw_name)
        if not canonical:
            unmatched_products.add(raw_name)

        size = parse_size(size_val, raw_name)
        color = str(color_val).strip() if color_val else None

        item = models.OrderItem(
            order_id=order.id,
            product_name=canonical or raw_name,
            size=size,
            color=color,
            quantity=qty,
        )
        db.add(item)

        existing_caleo_ids.add(caleo_id)
        created += 1

    db.commit()

    return {
        "created_orders": created,
        "skipped_orders": skipped,
        "products_created": len([p for p in PRODUCTS if p["name"] not in existing_products]),
        "unmatched_product_names": len(unmatched_products),
        "unmatched_samples": list(unmatched_products)[:20],
    }
